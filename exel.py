import openpyxl
excel_document = openpyxl.load_workbook('sample.xlsx')
print(type(excel_document))
print("Hojas en el documento:", excel_document.sheetnames)
sheet = excel_document['Sheet1']
print("Valor en la celda A2:", sheet['A2'].value)
print("Valor en fila 5, columna 2:", sheet.cell(row=5, column=2).value)
print("Valores en el rango A1:B3:")
for row in sheet['A1':'B3']:
    for cell in row:
        print(cell.value)
print("Valores en todas las filas:")
for row in sheet.iter_rows():
    for cell in row:
        print(cell.value)
print("Valores en todas las columnas:")
for col in sheet.iter_cols():
    for cell in col:
        print(cell.value)
